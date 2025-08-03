from typing import List, Dict, Any, Optional
from datetime import datetime
import pandas as pd
import numpy as np
from sqlalchemy.orm import Session
from sqlalchemy import and_, func
import logging
import os
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from ..models import Resource, Project, ResourceAllocation, CapacityPlan, Report
from ..schemas import ReportGenerationRequest, Report as ReportSchema

logger = logging.getLogger(__name__)

class ReportService:
    def __init__(self, db: Session):
        self.db = db
        self.reports_dir = "reports"
        os.makedirs(self.reports_dir, exist_ok=True)
    
    def generate_monthly_report(self, request: ReportGenerationRequest) -> ReportSchema:
        """Generate monthly capacity planning report"""
        try:
            # Create report record
            report = Report(
                title=f"Monthly Capacity Report - {request.period_start.strftime('%B %Y')}",
                type=request.report_type.value,
                period_start=request.period_start,
                period_end=request.period_end,
                format=request.format.value,
                status="generating",
                parameters=self._serialize_parameters(request),
                generated_by="system"
            )
            self.db.add(report)
            self.db.commit()
            
            # Generate report content
            report_data = self._generate_report_data(request)
            
            # Create file based on format
            if request.format.value == "pdf":
                file_path = self._generate_pdf_report(report_data, report.id)
            elif request.format.value == "excel":
                file_path = self._generate_excel_report(report_data, report.id)
            else:
                file_path = self._generate_html_report(report_data, report.id)
            
            # Update report record
            report.file_path = file_path
            report.status = "completed"
            report.completed_at = datetime.now()
            self.db.commit()
            
            return ReportSchema.from_orm(report)
            
        except Exception as e:
            logger.error(f"Error generating report: {e}")
            if 'report' in locals():
                report.status = "failed"
                self.db.commit()
            raise
    
    def _serialize_parameters(self, request: ReportGenerationRequest) -> Dict[str, Any]:
        """Serialize request parameters for database storage"""
        params = request.dict()
        # Convert datetime objects to ISO format strings
        for key, value in params.items():
            if isinstance(value, datetime):
                params[key] = value.isoformat()
        return params
    
    def _generate_report_data(self, request: ReportGenerationRequest) -> Dict[str, Any]:
        """Generate data for the report"""
        # Get capacity plans for the period
        plans = self.db.query(CapacityPlan).filter(
            and_(
                CapacityPlan.month == request.period_start.month,
                CapacityPlan.year == request.period_start.year
            )
        ).all()
        
        # Get resources and projects
        resources = self.db.query(Resource).all()
        projects = self.db.query(Project).all()
        
        # Calculate summary statistics
        total_planned_capacity = sum(plan.planned_capacity for plan in plans)
        total_actual_capacity = sum(plan.actual_capacity for plan in plans)
        avg_utilization = sum(plan.utilization_rate for plan in plans) / len(plans) if plans else 0
        avg_efficiency = sum(plan.efficiency_score for plan in plans) / len(plans) if plans else 0
        
        # Resource utilization breakdown
        resource_utilization = {}
        for resource in resources:
            resource_plans = [p for p in plans if p.resource_id == resource.id]
            if resource_plans:
                resource_utilization[resource.name] = {
                    'total_planned': sum(p.planned_capacity for p in resource_plans),
                    'total_actual': sum(p.actual_capacity for p in resource_plans),
                    'avg_utilization': sum(p.utilization_rate for p in resource_plans) / len(resource_plans),
                    'avg_efficiency': sum(p.efficiency_score for p in resource_plans) / len(resource_plans)
                }
        
        # Project allocation breakdown
        project_allocation = {}
        for project in projects:
            project_plans = [p for p in plans if p.project_id == project.id]
            if project_plans:
                project_allocation[project.name] = {
                    'total_allocated': sum(p.planned_capacity for p in project_plans),
                    'resources_count': len(set(p.resource_id for p in project_plans)),
                    'avg_efficiency': sum(p.efficiency_score for p in project_plans) / len(project_plans)
                }
        
        return {
            'period': f"{request.period_start.strftime('%B %Y')}",
            'summary': {
                'total_planned_capacity': total_planned_capacity,
                'total_actual_capacity': total_actual_capacity,
                'average_utilization': avg_utilization,
                'average_efficiency': avg_efficiency,
                'total_plans': len(plans),
                'total_resources': len(resources),
                'total_projects': len(projects)
            },
            'resource_utilization': resource_utilization,
            'project_allocation': project_allocation,
            'plans': [
                {
                    'resource_name': next((r.name for r in resources if r.id == p.resource_id), 'Unknown'),
                    'project_name': next((pr.name for pr in projects if pr.id == p.project_id), 'Unknown'),
                    'planned_capacity': p.planned_capacity,
                    'actual_capacity': p.actual_capacity,
                    'utilization_rate': p.utilization_rate,
                    'efficiency_score': p.efficiency_score
                }
                for p in plans
            ]
        }
    
    def _generate_pdf_report(self, data: Dict[str, Any], report_id: int) -> str:
        """Generate PDF report"""
        filename = f"capacity_report_{report_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join(self.reports_dir, filename)
        
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        story.append(Paragraph(f"Capacity Planning Report - {data['period']}", title_style))
        story.append(Spacer(1, 20))
        
        # Summary section
        story.append(Paragraph("Summary Statistics", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Planned Capacity', f"{data['summary']['total_planned_capacity']:.2f} hours"],
            ['Total Actual Capacity', f"{data['summary']['total_actual_capacity']:.2f} hours"],
            ['Average Utilization', f"{data['summary']['average_utilization']:.2%}"],
            ['Average Efficiency', f"{data['summary']['average_efficiency']:.2%}"],
            ['Total Plans', str(data['summary']['total_plans'])],
            ['Total Resources', str(data['summary']['total_resources'])],
            ['Total Projects', str(data['summary']['total_projects'])]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Resource utilization section
        if data['resource_utilization']:
            story.append(Paragraph("Resource Utilization", styles['Heading2']))
            story.append(Spacer(1, 12))
            
            resource_data = [['Resource', 'Planned (hrs)', 'Actual (hrs)', 'Utilization', 'Efficiency']]
            for resource, stats in data['resource_utilization'].items():
                resource_data.append([
                    resource,
                    f"{stats['total_planned']:.2f}",
                    f"{stats['total_actual']:.2f}",
                    f"{stats['avg_utilization']:.2%}",
                    f"{stats['avg_efficiency']:.2%}"
                ])
            
            resource_table = Table(resource_data, colWidths=[2*inch, 1*inch, 1*inch, 1*inch, 1*inch])
            resource_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(resource_table)
            story.append(Spacer(1, 20))
        
        # Build PDF
        doc.build(story)
        return filepath
    
    def _generate_excel_report(self, data: Dict[str, Any], report_id: int) -> str:
        """Generate Excel report"""
        filename = f"capacity_report_{report_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.reports_dir, filename)
        
        wb = openpyxl.Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Add summary data
        ws_summary['A1'] = f"Capacity Planning Report - {data['period']}"
        ws_summary['A1'].font = Font(size=16, bold=True)
        
        summary_headers = ['Metric', 'Value']
        for i, header in enumerate(summary_headers, 1):
            ws_summary[f'A{i+2}'] = header
            ws_summary[f'A{i+2}'].font = Font(bold=True)
        
        summary_data = [
            ['Total Planned Capacity', f"{data['summary']['total_planned_capacity']:.2f} hours"],
            ['Total Actual Capacity', f"{data['summary']['total_actual_capacity']:.2f} hours"],
            ['Average Utilization', f"{data['summary']['average_utilization']:.2%}"],
            ['Average Efficiency', f"{data['summary']['average_efficiency']:.2%}"],
            ['Total Plans', str(data['summary']['total_plans'])],
            ['Total Resources', str(data['summary']['total_resources'])],
            ['Total Projects', str(data['summary']['total_projects'])]
        ]
        
        for i, row in enumerate(summary_data, 3):
            for j, value in enumerate(row, 1):
                ws_summary[f'{chr(64+j)}{i}'] = value
        
        # Resource utilization sheet
        if data['resource_utilization']:
            ws_resources = wb.create_sheet("Resource Utilization")
            
            headers = ['Resource', 'Planned (hrs)', 'Actual (hrs)', 'Utilization', 'Efficiency']
            for i, header in enumerate(headers, 1):
                ws_resources[f'{chr(64+i)}1'] = header
                ws_resources[f'{chr(64+i)}1'].font = Font(bold=True)
            
            for i, (resource, stats) in enumerate(data['resource_utilization'].items(), 2):
                ws_resources[f'A{i}'] = resource
                ws_resources[f'B{i}'] = stats['total_planned']
                ws_resources[f'C{i}'] = stats['total_actual']
                ws_resources[f'D{i}'] = stats['avg_utilization']
                ws_resources[f'E{i}'] = stats['avg_efficiency']
        
        # Project allocation sheet
        if data['project_allocation']:
            ws_projects = wb.create_sheet("Project Allocation")
            
            headers = ['Project', 'Total Allocated (hrs)', 'Resources Count', 'Avg Efficiency']
            for i, header in enumerate(headers, 1):
                ws_projects[f'{chr(64+i)}1'] = header
                ws_projects[f'{chr(64+i)}1'].font = Font(bold=True)
            
            for i, (project, stats) in enumerate(data['project_allocation'].items(), 2):
                ws_projects[f'A{i}'] = project
                ws_projects[f'B{i}'] = stats['total_allocated']
                ws_projects[f'C{i}'] = stats['resources_count']
                ws_projects[f'D{i}'] = stats['avg_efficiency']
        
        # Plans detail sheet
        if data['plans']:
            ws_plans = wb.create_sheet("Capacity Plans")
            
            headers = ['Resource', 'Project', 'Planned Capacity', 'Actual Capacity', 'Utilization', 'Efficiency']
            for i, header in enumerate(headers, 1):
                ws_plans[f'{chr(64+i)}1'] = header
                ws_plans[f'{chr(64+i)}1'].font = Font(bold=True)
            
            for i, plan in enumerate(data['plans'], 2):
                ws_plans[f'A{i}'] = plan['resource_name']
                ws_plans[f'B{i}'] = plan['project_name']
                ws_plans[f'C{i}'] = plan['planned_capacity']
                ws_plans[f'D{i}'] = plan['actual_capacity']
                ws_plans[f'E{i}'] = plan['utilization_rate']
                ws_plans[f'F{i}'] = plan['efficiency_score']
        
        wb.save(filepath)
        return filepath
    
    def _generate_html_report(self, data: Dict[str, Any], report_id: int) -> str:
        """Generate HTML report"""
        filename = f"capacity_report_{report_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
        filepath = os.path.join(self.reports_dir, filename)
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Capacity Planning Report - {data['period']}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #333; text-align: center; }}
                h2 {{ color: #666; border-bottom: 2px solid #ddd; }}
                table {{ border-collapse: collapse; width: 100%; margin: 10px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; font-weight: bold; }}
                tr:nth-child(even) {{ background-color: #f9f9f9; }}
                .summary {{ background-color: #e7f3ff; padding: 15px; border-radius: 5px; }}
            </style>
        </head>
        <body>
            <h1>Capacity Planning Report - {data['period']}</h1>
            
            <h2>Summary Statistics</h2>
            <div class="summary">
                <p><strong>Total Planned Capacity:</strong> {data['summary']['total_planned_capacity']:.2f} hours</p>
                <p><strong>Total Actual Capacity:</strong> {data['summary']['total_actual_capacity']:.2f} hours</p>
                <p><strong>Average Utilization:</strong> {data['summary']['average_utilization']:.2%}</p>
                <p><strong>Average Efficiency:</strong> {data['summary']['average_efficiency']:.2%}</p>
                <p><strong>Total Plans:</strong> {data['summary']['total_plans']}</p>
                <p><strong>Total Resources:</strong> {data['summary']['total_resources']}</p>
                <p><strong>Total Projects:</strong> {data['summary']['total_projects']}</p>
            </div>
        """
        
        if data['resource_utilization']:
            html_content += """
            <h2>Resource Utilization</h2>
            <table>
                <tr>
                    <th>Resource</th>
                    <th>Planned (hrs)</th>
                    <th>Actual (hrs)</th>
                    <th>Utilization</th>
                    <th>Efficiency</th>
                </tr>
            """
            for resource, stats in data['resource_utilization'].items():
                html_content += f"""
                <tr>
                    <td>{resource}</td>
                    <td>{stats['total_planned']:.2f}</td>
                    <td>{stats['total_actual']:.2f}</td>
                    <td>{stats['avg_utilization']:.2%}</td>
                    <td>{stats['avg_efficiency']:.2%}</td>
                </tr>
                """
            html_content += "</table>"
        
        if data['project_allocation']:
            html_content += """
            <h2>Project Allocation</h2>
            <table>
                <tr>
                    <th>Project</th>
                    <th>Total Allocated (hrs)</th>
                    <th>Resources Count</th>
                    <th>Avg Efficiency</th>
                </tr>
            """
            for project, stats in data['project_allocation'].items():
                html_content += f"""
                <tr>
                    <td>{project}</td>
                    <td>{stats['total_allocated']:.2f}</td>
                    <td>{stats['resources_count']}</td>
                    <td>{stats['avg_efficiency']:.2%}</td>
                </tr>
                """
            html_content += "</table>"
        
        html_content += """
        </body>
        </html>
        """
        
        with open(filepath, 'w') as f:
            f.write(html_content)
        
        return filepath
    
    def get_reports(self, skip: int = 0, limit: int = 100) -> List[ReportSchema]:
        """Get list of reports"""
        reports = self.db.query(Report).offset(skip).limit(limit).all()
        return [ReportSchema.from_orm(report) for report in reports]
    
    def get_report(self, report_id: int) -> Optional[ReportSchema]:
        """Get specific report by ID"""
        report = self.db.query(Report).filter(Report.id == report_id).first()
        return ReportSchema.from_orm(report) if report else None 